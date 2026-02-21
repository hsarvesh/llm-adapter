"""Parser for spreadsheet formats: XLSX."""

import io
import structlog

from parsers.base import BaseParser

logger = structlog.get_logger(__name__)


class SpreadsheetParser(BaseParser):
    """Handles Excel spreadsheet formats."""

    @property
    def supported_extensions(self) -> list[str]:
        return [".xlsx", ".xls"]

    def parse(self, file_bytes: bytes, filename: str) -> str:
        """Extract content from Excel files using openpyxl."""
        try:
            from openpyxl import load_workbook

            wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
            sheets_text = []

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = []

                for row in ws.iter_rows(values_only=True):
                    cells = [str(cell) if cell is not None else "" for cell in row]
                    # Skip completely empty rows
                    if any(cell.strip() for cell in cells):
                        rows.append(cells)

                if not rows:
                    sheets_text.append(f"--- Sheet: {sheet_name} ---\n(Empty sheet)")
                    continue

                # Calculate column widths for formatting
                col_count = max(len(row) for row in rows)
                col_widths = [0] * col_count
                for row in rows:
                    for i, cell in enumerate(row):
                        if i < col_count:
                            col_widths[i] = max(col_widths[i], min(len(cell), 40))

                # Format as table
                lines = []
                for row_idx, row in enumerate(rows):
                    padded = []
                    for i in range(col_count):
                        cell = row[i] if i < len(row) else ""
                        width = col_widths[i] if i < len(col_widths) else 10
                        padded.append(cell[:40].ljust(width))
                    lines.append(" | ".join(padded))
                    if row_idx == 0:
                        separator = "-+-".join("-" * w for w in col_widths)
                        lines.append(separator)

                sheet_content = "\n".join(lines)
                sheets_text.append(f"--- Sheet: {sheet_name} ({len(rows)} rows) ---\n{sheet_content}")

            wb.close()

            result = "\n\n".join(sheets_text)
            logger.info("xlsx_parsed", filename=filename, sheets=len(sheets_text))
            return result
        except Exception as e:
            logger.error("xlsx_parse_error", filename=filename, error=str(e))
            return f"[Error extracting spreadsheet content: {str(e)}]"
